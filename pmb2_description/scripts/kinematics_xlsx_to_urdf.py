#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""
Created on 04/02/15

@author: Enrique Fernandez

Converts the kinematics specification from a xlsx file into URDF

pip install openpyxl
"""

from openpyxl import load_workbook
import sys

(JOINT_COLUMN, CHILD_LINK_COLUMN, PARENT_LINK_COLUMN, JOINT_TYPE_COLUMN, COMMENT_COLUMN, MASS_COLUMN, TX_COLUMN, TY_COLUMN, TZ_COLUMN, RX_COLUMN, RY_COLUMN, RZ_COLUMN, CX_COLUMN, CY_COLUMN, CZ_COLUMN, LXX_COLUMN, LXY_COLUMN, LXZ_COLUMN, LYY_COLUMN, LYZ_COLUMN, LZZ_COLUMN) = range(3, 24)

def convert(filename):
    wb = load_workbook(filename, data_only=True)

    sheet_names = wb.get_sheet_names()
    sheet = wb[sheet_names[0]]

    try:
        row = 10
        while True:
            # Joint
            joint = sheet.cell(row=row, column=JOINT_COLUMN).value
            if joint is None:
                break
            joint = str(joint)

            # Child and parent link
            child_link = str(sheet.cell(row=row, column=CHILD_LINK_COLUMN).value)
            parent_link = str(sheet.cell(row=row, column=PARENT_LINK_COLUMN).value)

            # Type
            joint_type = str(sheet.cell(row=row, column=JOINT_TYPE_COLUMN).value)

            # Mass
            mass = 1e-3 * float(sheet.cell(row=row, column=MASS_COLUMN).value)

            # Frame transformations: translation
            tx = 1e-3 * float(sheet.cell(row=row, column=TX_COLUMN).value)
            ty = 1e-3 * float(sheet.cell(row=row, column=TY_COLUMN).value)
            tz = 1e-3 * float(sheet.cell(row=row, column=TZ_COLUMN).value)

            # Frame transformations: rotation
            rx = float(sheet.cell(row=row, column=RX_COLUMN).value)
            ry = float(sheet.cell(row=row, column=RY_COLUMN).value)
            rz = float(sheet.cell(row=row, column=RZ_COLUMN).value)

            # CoM location wrt link
            cx = 1e-3 * float(sheet.cell(row=row, column=CX_COLUMN).value)
            cy = 1e-3 * float(sheet.cell(row=row, column=CY_COLUMN).value)
            cz = 1e-3 * float(sheet.cell(row=row, column=CZ_COLUMN).value)

            # Moments of Inertia wrt CoM and aligned with link
            lxx = 1e-9 * float(sheet.cell(row=row, column=LXX_COLUMN).value)
            lxy = 1e-9 * float(sheet.cell(row=row, column=LXY_COLUMN).value)
            lxz = 1e-9 * float(sheet.cell(row=row, column=LXZ_COLUMN).value)
            lyy = 1e-9 * float(sheet.cell(row=row, column=LYY_COLUMN).value)
            lyz = 1e-9 * float(sheet.cell(row=row, column=LYZ_COLUMN).value)
            lzz = 1e-9 * float(sheet.cell(row=row, column=LZZ_COLUMN).value)

            # URDF link
            print '''
            <!-- Joint %s : %s -> %s -->
            ''' % (joint, child_link, parent_link)
            print '''
            <link name="%s">
              <inertial>
                <origin xyz="%0.11f %0.11f %0.11f" rpy="%0.11f %0.11f %0.11f"/>
                <mass value="%0.11f"/>
                <inertia ixx="%0.11f" ixy="%0.11f" ixz="%0.11f"
                         iyy="%0.11f" iyz="%0.11f"
                         izz="%0.11f"/>
              </inertial>
            </link>''' % (child_link, cx, cy, cz, 0, 0, 0, mass, lxx, lxy, lxz, lyy, lyz, lzz)

            # URDF joint for parent link to link
            print '''
            <joint name="%s" type="%s">
              <origin xyz="%0.11f %0.11f %0.11f" rpy="${%0.11f * deg_to_rad} ${%0.11f * deg_to_rad} ${%0.11f * deg_to_rad}"/>
              <child link="%s"/>
              <parent link="%s"/>
            </joint>''' % (joint, joint_type, tx, ty, tz, rx, ry, rz, child_link, parent_link)

            row = row + 1

    except Exception as e:
        print 'Error reading cells: %s' % e

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print "No xlsx file given!"
        exit(0)

    convert(sys.argv[1])

